import discord
import asyncio
import openpyxl
import os

client = discord.Client()

@client.event
async def on_ready():
    print(client.user.id)
    print("준비 완료 했습니다. 보스")
    print("--------------------")
    game = discord.Game("Madby:김조코")
    await client.change_presence(status=discord.Status.online, activity=game)

@client.event
async def on_message(message):
    if message.content.startswith("조미야"):
        file = openpyxl.load_workbook("학습데이터베이스.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1,3000):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break

    if message.content.startswith("조미학습") and not message.content.startswith("조미학습제거"):
        file = openpyxl.load_workbook("학습데이터베이스.xlsx")
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 3000):
            if sheet["A" + str(i)].value == "-" or sheet["A" + str(i)].value == learn[1]:
                sheet["A" + str(i)].value = learn[1]
                sheet["A" + str(i+1)].value = '-'
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("학습완료.")
                print("학습완료.", learn[1],learn[2])
                break
        file.save("학습데이터베이스.xlsx")

    if message.content.startswith("조미학습제거"):
        file = openpyxl.load_workbook("학습데이터베이스.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 3000):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = '-'
                sheet["A" + str(i+1)].value = ''
                sheet["B" + str(i)].value = ''
                await message.channel.send(memory[1] + " 을(를) 잊엇습니다.")
                print (memory[1] + " 를(을) 잊었습니다.")
        file.save("학습데이터베이스.xlsx")
    
access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
